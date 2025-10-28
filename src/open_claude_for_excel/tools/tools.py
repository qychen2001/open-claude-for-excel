import logging
import os
from typing import Any, Dict, List, Optional

# from mcp.server.fastmcp import FastMCP
from langchain_core.tools import tool

from open_claude_for_excel.tools.chart import create_chart_in_sheet as create_chart_impl
from open_claude_for_excel.tools.data import write_data

# Import exceptions
from open_claude_for_excel.tools.exceptions import (
    CalculationError,
    ChartError,
    DataError,
    FormattingError,
    PivotError,
    SheetError,
    ValidationError,
    WorkbookError,
)
from open_claude_for_excel.tools.pivot import (
    create_pivot_table as create_pivot_table_impl,
)
from open_claude_for_excel.tools.sheet import (
    copy_sheet,
    delete_cols,
    delete_rows,
    delete_sheet,
    get_merged_ranges,
    insert_cols,
    insert_row,
    merge_range,
    rename_sheet,
    unmerge_range,
)
from open_claude_for_excel.tools.tables import create_excel_table as create_table_impl

# Import from open_claude_for_excel.tools package with consistent _impl suffixes
from open_claude_for_excel.tools.validation import (
    validate_formula_in_cell_operation as validate_formula_impl,
)
from open_claude_for_excel.tools.validation import (
    validate_range_in_sheet_operation as validate_range_impl,
)
from open_claude_for_excel.tools.workbook import get_workbook_info

# Get project root directory path for log file path.
# When using the stdio transmission method,
# relative paths may cause log files to fail to create
# due to the client's running location and permission issues,
# resulting in the program not being able to run.
# Thus using os.path.join(ROOT_DIR, "excel-mcp.log") instead.

ROOT_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
LOG_FILE = os.path.join(ROOT_DIR, "open-claude-for-excel.log")

# Initialize EXCEL_FILES_PATH variable without assigning a value
OPEN_CLAUDE_FOR_EXCEL_FILES_PATH = None

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    handlers=[
        # Referring to https://github.com/modelcontextprotocol/python-sdk/issues/409#issuecomment-2816831318
        # The stdio mode server MUST NOT write anything to its stdout that is not a valid MCP message.
        logging.FileHandler(LOG_FILE)
    ],
)
logger = logging.getLogger("excel-mcp")
# # Initialize FastMCP server
# mcp = FastMCP(
#     "excel-mcp",
#     host=os.environ.get("FASTMCP_HOST", "0.0.0.0"),
#     port=int(os.environ.get("FASTMCP_PORT", "8017")),
#     instructions="Excel MCP Server for manipulating Excel files",
# )


def get_excel_path(filename: str) -> str:
    """Get full path to Excel file.

    Args:
        filename: Name of Excel file

    Returns:
        Full path to Excel file
    """
    # If filename is already an absolute path, return it
    if os.path.isabs(filename):
        return filename

    # Check if in SSE mode (EXCEL_FILES_PATH is not None)
    if OPEN_CLAUDE_FOR_EXCEL_FILES_PATH is None:
        # Must use absolute path
        raise ValueError(
            f"Invalid filename: {filename}, must be an absolute path when not in SSE mode"
        )

    # In SSE mode, if it's a relative path, resolve it based on EXCEL_FILES_PATH
    return os.path.join(OPEN_CLAUDE_FOR_EXCEL_FILES_PATH, filename)


@tool(parse_docstring=True)
def apply_formula(
    filepath: str,
    sheet_name: str,
    cell: str,
    formula: str,
) -> str:
    """Apply Excel formula to cell. Excel formula will write to cell with verification.

    Args:
        filepath: Path to Excel file
        sheet_name: Target worksheet name
        cell: Target cell reference
        formula: Excel formula to apply

    Returns:
        Success message
    """
    try:
        full_path = get_excel_path(filepath)
        # First validate the formula
        validation = validate_formula_impl(full_path, sheet_name, cell, formula)
        if isinstance(validation, dict) and "error" in validation:
            return f"Error: {validation['error']}"

        # If valid, apply the formula
        from open_claude_for_excel.tools.calculations import (
            apply_formula as apply_formula_impl,
        )

        result = apply_formula_impl(full_path, sheet_name, cell, formula)
        return result["message"]
    except (ValidationError, CalculationError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error applying formula: {e}")
        raise


@tool(parse_docstring=True)
def validate_formula_syntax(
    filepath: str,
    sheet_name: str,
    cell: str,
    formula: str,
) -> str:
    """Validate Excel formula syntax without applying it.

    Args:
        filepath: Path to Excel file
        sheet_name: Target worksheet name
        cell: Target cell reference
        formula: Excel formula to validate

    Returns:
        Validation result message
    """
    try:
        full_path = get_excel_path(filepath)
        result = validate_formula_impl(full_path, sheet_name, cell, formula)
        return result["message"]
    except (ValidationError, CalculationError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error validating formula: {e}")
        raise


@tool(parse_docstring=True)
def format_range(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: Optional[str] = None,
    bold: bool = False,
    italic: bool = False,
    underline: bool = False,
    font_size: Optional[int] = None,
    font_color: Optional[str] = None,
    bg_color: Optional[str] = None,
    border_style: Optional[str] = None,
    border_color: Optional[str] = None,
    number_format: Optional[str] = None,
    alignment: Optional[str] = None,
    wrap_text: bool = False,
    merge_cells: bool = False,
    protection: Optional[Dict[str, Any]] = None,
    conditional_format: Optional[Dict[str, Any]] = None,
) -> str:
    """Apply formatting to a range of cells.

    Args:
        filepath: Path to Excel file
        sheet_name: Target worksheet name
        start_cell: Starting cell reference
        end_cell: Optional ending cell reference
        bold: Whether to make text bold
        italic: Whether to make text italic
        underline: Whether to underline text
        font_size: Font size in points
        font_color: Font color (hex code)
        bg_color: Background color (hex code)
        border_style: Border style (thin, medium, thick, double)
        border_color: Border color (hex code)
        number_format: Excel number format string
        alignment: Text alignment (left, center, right, justify)
        wrap_text: Whether to wrap text
        merge_cells: Whether to merge the range
        protection: Cell protection settings
        conditional_format: Conditional formatting rules

    Returns:
        Success message
    """
    try:
        full_path = get_excel_path(filepath)
        from open_claude_for_excel.tools.formatting import (
            format_range as format_range_func,
        )

        # Convert None values to appropriate defaults for the underlying function
        format_range_func(
            filepath=full_path,
            sheet_name=sheet_name,
            start_cell=start_cell,
            end_cell=end_cell,  # This can be None
            bold=bold,
            italic=italic,
            underline=underline,
            font_size=font_size,  # This can be None
            font_color=font_color,  # This can be None
            bg_color=bg_color,  # This can be None
            border_style=border_style,  # This can be None
            border_color=border_color,  # This can be None
            number_format=number_format,  # This can be None
            alignment=alignment,  # This can be None
            wrap_text=wrap_text,
            merge_cells=merge_cells,
            protection=protection,  # This can be None
            conditional_format=conditional_format,  # This can be None
        )
        return "Range formatted successfully"
    except (ValidationError, FormattingError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error formatting range: {e}")
        raise


@tool(parse_docstring=True)
def read_data_from_excel(
    filepath: str,
    sheet_name: str,
    start_cell: str = "A1",
    end_cell: Optional[str] = None,
    preview_only: bool = False,
) -> str:
    """Read data from Excel worksheet with cell metadata including validation rules.

    Args:
        filepath: Path to Excel file
        sheet_name: Name of worksheet
        start_cell: Starting cell (default A1)
        end_cell: Ending cell (optional, auto-expands if not provided)
        preview_only: Whether to return preview only

    Returns:
        JSON string containing structured cell data with validation metadata. Each cell includes: address, value, row, column, and validation info (if any).
    """
    try:
        full_path = get_excel_path(filepath)
        from open_claude_for_excel.tools.data import read_excel_range_with_metadata

        result = read_excel_range_with_metadata(
            full_path, sheet_name, start_cell, end_cell
        )
        if not result or not result.get("cells"):
            return "No data found in specified range"

        # Return as formatted JSON string
        import json

        return json.dumps(result, indent=2, default=str)

    except Exception as e:
        logger.error(f"Error reading data: {e}")
        raise


@tool(parse_docstring=True)
def write_data_to_excel(
    filepath: str,
    sheet_name: str,
    data: List[List],
    start_cell: str = "A1",
) -> str:
    """Write data to Excel worksheet. Excel formula will write to cell without any verification.

    Args:
        filepath: Path to Excel file
        sheet_name: Target worksheet name
        data: List of lists containing data to write to the worksheet, sublists are assumed to be rows
        start_cell: Cell to start writing to, default is "A1"

    Returns:
        Success message
    """
    try:
        full_path = get_excel_path(filepath)
        result = write_data(full_path, sheet_name, data, start_cell)
        return result["message"]
    except (ValidationError, DataError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error writing data: {e}")
        raise


@tool(parse_docstring=True)
def create_workbook(filepath: str) -> str:
    """Create new Excel workbook.

    Args:
        filepath: Path where to create workbook

    Returns:
        Success message with created file path
    """
    try:
        full_path = get_excel_path(filepath)
        from open_claude_for_excel.tools.workbook import (
            create_workbook as create_workbook_impl,
        )

        create_workbook_impl(full_path)
        return f"Created workbook at {full_path}"
    except WorkbookError as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating workbook: {e}")
        raise


@tool(parse_docstring=True)
def create_worksheet(filepath: str, sheet_name: str) -> str:
    """Create new worksheet in workbook.

    Args:
        filepath: Path to Excel file
        sheet_name: Name for the new worksheet

    Returns:
        Success message
    """
    try:
        full_path = get_excel_path(filepath)
        from open_claude_for_excel.tools.workbook import (
            create_sheet as create_worksheet_impl,
        )

        result = create_worksheet_impl(full_path, sheet_name)
        return result["message"]
    except (ValidationError, WorkbookError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating worksheet: {e}")
        raise


@tool(parse_docstring=True)
def create_chart(
    filepath: str,
    sheet_name: str,
    data_range: str,
    chart_type: str,
    target_cell: str,
    title: str = "",
    x_axis: str = "",
    y_axis: str = "",
) -> str:
    """Create chart in worksheet.

    Args:
        filepath: Path to Excel file
        sheet_name: Target worksheet name
        data_range: Range containing chart data
        chart_type: Type of chart (line, bar, pie, scatter, area)
        target_cell: Cell where to place chart
        title: Optional chart title
        x_axis: Optional x-axis label
        y_axis: Optional y-axis label

    Returns:
        Success message
    """
    try:
        full_path = get_excel_path(filepath)
        result = create_chart_impl(
            filepath=full_path,
            sheet_name=sheet_name,
            data_range=data_range,
            chart_type=chart_type,
            target_cell=target_cell,
            title=title,
            x_axis=x_axis,
            y_axis=y_axis,
        )
        return result["message"]
    except (ValidationError, ChartError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating chart: {e}")
        raise


@tool(parse_docstring=True)
def create_pivot_table(
    filepath: str,
    sheet_name: str,
    data_range: str,
    rows: List[str],
    values: List[str],
    columns: Optional[List[str]] = None,
    agg_func: str = "mean",
) -> str:
    """Create pivot table in worksheet.

    Args:
        filepath: Path to Excel file
        sheet_name: Target worksheet name
        data_range: Range containing source data
        rows: Fields for row labels
        values: Fields for values
        columns: Optional fields for column labels
        agg_func: Aggregation function (sum, count, average, max, min)

    Returns:
        Success message
    """
    try:
        full_path = get_excel_path(filepath)
        result = create_pivot_table_impl(
            filepath=full_path,
            sheet_name=sheet_name,
            data_range=data_range,
            rows=rows,
            values=values,
            columns=columns or [],
            agg_func=agg_func,
        )
        return result["message"]
    except (ValidationError, PivotError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating pivot table: {e}")
        raise


@tool(parse_docstring=True)
def create_table(
    filepath: str,
    sheet_name: str,
    data_range: str,
    table_name: Optional[str] = None,
    table_style: str = "TableStyleMedium9",
) -> str:
    """Creates a native Excel table from a specified range of data.

    Args:
        filepath: Path to the Excel file.
        sheet_name: Name of the worksheet.
        data_range: The cell range for the table (e.g., "A1:D5").
        table_name: Optional unique name for the table.
        table_style: Optional visual style for the table.

    Returns:
        Success message
    """
    try:
        full_path = get_excel_path(filepath)
        result = create_table_impl(
            filepath=full_path,
            sheet_name=sheet_name,
            data_range=data_range,
            table_name=table_name,
            table_style=table_style,
        )
        return result["message"]
    except DataError as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error creating table: {e}")
        raise


@tool(parse_docstring=True)
def copy_worksheet(filepath: str, source_sheet: str, target_sheet: str) -> str:
    """Copy worksheet within workbook.

    Args:
        filepath: Path to Excel file
        source_sheet: Name of sheet to copy
        target_sheet: Name for new sheet

    Returns:
        Success message
    """
    try:
        full_path = get_excel_path(filepath)
        result = copy_sheet(full_path, source_sheet, target_sheet)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error copying worksheet: {e}")
        raise


@tool(parse_docstring=True)
def delete_worksheet(filepath: str, sheet_name: str) -> str:
    """Delete worksheet from workbook.

    Args:
        filepath: Path to Excel file
        sheet_name: Name of sheet to delete

    Returns:
        Success message
    """
    try:
        full_path = get_excel_path(filepath)
        result = delete_sheet(full_path, sheet_name)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error deleting worksheet: {e}")
        raise


@tool(parse_docstring=True)
def rename_worksheet(filepath: str, old_name: str, new_name: str) -> str:
    """Rename worksheet in workbook.

    Args:
        filepath: Path to Excel file
        old_name: Current sheet name
        new_name: New sheet name

    Returns:
        Success message
    """
    try:
        full_path = get_excel_path(filepath)
        result = rename_sheet(full_path, old_name, new_name)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error renaming worksheet: {e}")
        raise


@tool(parse_docstring=True)
def get_workbook_metadata(filepath: str, include_ranges: bool = False) -> str:
    """Get metadata about workbook including sheets, ranges, etc.

    Args:
        filepath: Path to Excel file
        include_ranges: Whether to include range information

    Returns:
        String representation of workbook metadata
    """
    try:
        full_path = get_excel_path(filepath)
        result = get_workbook_info(full_path, include_ranges=include_ranges)
        return str(result)
    except WorkbookError as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error getting workbook metadata: {e}")
        raise


@tool(parse_docstring=True)
def merge_cells(filepath: str, sheet_name: str, start_cell: str, end_cell: str) -> str:
    """Merge a range of cells.

    Args:
        filepath: Path to Excel file
        sheet_name: Target worksheet name
        start_cell: Starting cell of range
        end_cell: Ending cell of range

    Returns:
        Success message
    """
    try:
        full_path = get_excel_path(filepath)
        result = merge_range(full_path, sheet_name, start_cell, end_cell)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error merging cells: {e}")
        raise


@tool(parse_docstring=True)
def unmerge_cells(
    filepath: str, sheet_name: str, start_cell: str, end_cell: str
) -> str:
    """Unmerge a range of cells.

    Args:
        filepath: Path to Excel file
        sheet_name: Target worksheet name
        start_cell: Starting cell of range
        end_cell: Ending cell of range

    Returns:
        Success message
    """
    try:
        full_path = get_excel_path(filepath)
        result = unmerge_range(full_path, sheet_name, start_cell, end_cell)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error unmerging cells: {e}")
        raise


@tool(parse_docstring=True)
def get_merged_cells(filepath: str, sheet_name: str) -> str:
    """Get merged cells in a worksheet.

    Args:
        filepath: Path to Excel file
        sheet_name: Target worksheet name

    Returns:
        String representation of merged cells
    """
    try:
        full_path = get_excel_path(filepath)
        return str(get_merged_ranges(full_path, sheet_name))
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error getting merged cells: {e}")
        raise


@tool(parse_docstring=True)
def copy_range(
    filepath: str,
    sheet_name: str,
    source_start: str,
    source_end: str,
    target_start: str,
    target_sheet: Optional[str] = None,
) -> str:
    """Copy a range of cells to another location.

    Args:
        filepath: Path to Excel file
        sheet_name: Source worksheet name
        source_start: Starting cell of source range
        source_end: Ending cell of source range
        target_start: Starting cell for paste
        target_sheet: Optional target worksheet name

    Returns:
        Success message
    """
    try:
        full_path = get_excel_path(filepath)
        from open_claude_for_excel.tools.sheet import copy_range_operation

        result = copy_range_operation(
            full_path,
            sheet_name,
            source_start,
            source_end,
            target_start,
            target_sheet or sheet_name,  # Use source sheet if target_sheet is None
        )
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error copying range: {e}")
        raise


@tool(parse_docstring=True)
def delete_range(
    filepath: str,
    sheet_name: str,
    start_cell: str,
    end_cell: str,
    shift_direction: str = "up",
) -> str:
    """Delete a range of cells and shift remaining cells.

    Args:
        filepath: Path to Excel file
        sheet_name: Target worksheet name
        start_cell: Starting cell of range
        end_cell: Ending cell of range
        shift_direction: Direction to shift cells ("up" or "left")

    Returns:
        Success message
    """
    try:
        full_path = get_excel_path(filepath)
        from open_claude_for_excel.tools.sheet import delete_range_operation

        result = delete_range_operation(
            full_path, sheet_name, start_cell, end_cell, shift_direction
        )
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error deleting range: {e}")
        raise


@tool(parse_docstring=True)
def validate_excel_range(
    filepath: str, sheet_name: str, start_cell: str, end_cell: Optional[str] = None
) -> str:
    """Validate if a range exists and is properly formatted.

    Args:
        filepath: Path to Excel file
        sheet_name: Target worksheet name
        start_cell: Starting cell of range
        end_cell: Optional ending cell of range

    Returns:
        Success message
    """
    try:
        full_path = get_excel_path(filepath)
        range_str = start_cell if not end_cell else f"{start_cell}:{end_cell}"
        result = validate_range_impl(full_path, sheet_name, range_str)
        return result["message"]
    except ValidationError as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error validating range: {e}")
        raise


@tool(parse_docstring=True)
def get_data_validation_info(filepath: str, sheet_name: str) -> str:
    """Get all data validation rules in a worksheet. This tool helps identify which cell ranges have validation rules and what types of validation are applied.

    Args:
        filepath: Path to Excel file
        sheet_name: Name of worksheet

    Returns:
        JSON string containing all validation rules in the worksheet
    """
    try:
        full_path = get_excel_path(filepath)
        from openpyxl import load_workbook

        from open_claude_for_excel.tools.cell_validation import (
            get_all_validation_ranges,
        )

        wb = load_workbook(full_path, read_only=False)
        if sheet_name not in wb.sheetnames:
            return f"Error: Sheet '{sheet_name}' not found"

        ws = wb[sheet_name]
        validations = get_all_validation_ranges(ws)
        wb.close()

        if not validations:
            return "No data validation rules found in this worksheet"

        import json

        return json.dumps(
            {"sheet_name": sheet_name, "validation_rules": validations},
            indent=2,
            default=str,
        )

    except Exception as e:
        logger.error(f"Error getting validation info: {e}")
        raise


@tool(parse_docstring=True)
def insert_rows(filepath: str, sheet_name: str, start_row: int, count: int = 1) -> str:
    """Insert one or more rows starting at the specified row.

    Args:
        filepath: Path to Excel file
        sheet_name: Target worksheet name
        start_row: Row number where to start inserting (1-based)
        count: Number of rows to insert (default: 1)

    Returns:
        Success message
    """
    try:
        full_path = get_excel_path(filepath)
        result = insert_row(full_path, sheet_name, start_row, count)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error inserting rows: {e}")
        raise


@tool(parse_docstring=True)
def insert_columns(
    filepath: str, sheet_name: str, start_col: int, count: int = 1
) -> str:
    """Insert one or more columns starting at the specified column.

    Args:
        filepath: Path to Excel file
        sheet_name: Target worksheet name
        start_col: Column number where to start inserting (1-based)
        count: Number of columns to insert (default: 1)

    Returns:
        Success message
    """
    try:
        full_path = get_excel_path(filepath)
        result = insert_cols(full_path, sheet_name, start_col, count)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error inserting columns: {e}")
        raise


@tool(parse_docstring=True)
def delete_sheet_rows(
    filepath: str, sheet_name: str, start_row: int, count: int = 1
) -> str:
    """Delete one or more rows starting at the specified row.

    Args:
        filepath: Path to Excel file
        sheet_name: Target worksheet name
        start_row: Row number where to start deleting (1-based)
        count: Number of rows to delete (default: 1)

    Returns:
        Success message
    """
    try:
        full_path = get_excel_path(filepath)
        result = delete_rows(full_path, sheet_name, start_row, count)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error deleting rows: {e}")
        raise


@tool(parse_docstring=True)
def delete_sheet_columns(
    filepath: str, sheet_name: str, start_col: int, count: int = 1
) -> str:
    """Delete one or more columns starting at the specified column.

    Args:
        filepath: Path to Excel file
        sheet_name: Target worksheet name
        start_col: Column number where to start deleting (1-based)
        count: Number of columns to delete (default: 1)

    Returns:
        Success message
    """
    try:
        full_path = get_excel_path(filepath)
        result = delete_cols(full_path, sheet_name, start_col, count)
        return result["message"]
    except (ValidationError, SheetError) as e:
        return f"Error: {str(e)}"
    except Exception as e:
        logger.error(f"Error deleting columns: {e}")
        raise
